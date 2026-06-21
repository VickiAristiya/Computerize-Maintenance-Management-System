"""
Export hasil prediksi model ML ke Excel.
Jalankan: python export_excel.py
Output: hasil_prediksi_model.xlsx
"""
import sys, os, warnings, random
warnings.filterwarnings("ignore")
os.environ["TF_CPP_MIN_LOG_LEVEL"] = "3"

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "cmms-backend"))
from app.ml_service import CompressorPredictor, COMPONENT_NAME_ID, COMPONENTS

random.seed(42)
np.random.seed(42)

predictor = CompressorPredictor()
if not predictor.is_ready():
    print("ERROR: model tidak siap")
    sys.exit(1)

FEAT = predictor.feature_columns
COMP_NAMES = {c: COMPONENT_NAME_ID[c] for c in COMPONENTS}

# ── Load dataset ───────────────────────────────────────────────────────────────
CSV_PATH = os.path.join(os.path.dirname(__file__),
                        "machine-learning filtered data", "data.csv")
df_raw = pd.read_csv(CSV_PATH)

sensor_stats = {}
for f in FEAT:
    if f in df_raw.columns:
        col = df_raw[f].dropna()
        sensor_stats[f] = {
            "min": col.min(), "p10": col.quantile(0.10),
            "p25": col.quantile(0.25), "p50": col.quantile(0.50),
            "p75": col.quantile(0.75), "p90": col.quantile(0.90),
            "max": col.max(),
        }

# ── Predict helper ─────────────────────────────────────────────────────────────
def predict_row(sensors):
    result = predictor.predict(sensors)
    return result if result.get("ok") else None

def get_fp(result, comp):
    return result["components"][comp]["failure_probability"]

def get_hs(result, comp):
    return result["components"][comp]["health_score"]

def get_rl(result, comp):
    return result["components"][comp]["risk_level"]

def get_pred(result, comp):
    return result["components"][comp]["prediction"]

def get_status(result, comp):
    return result["components"][comp]["status"]

# ── Kumpulkan semua skenario ──────────────────────────────────────────────────
print("Mengumpulkan data...")

all_rows = []

def add_row(label, sensors):
    r = predict_row(sensors)
    if r is None:
        return
    row = {
        "skenario": label,
        "sumber": "",
    }
    for f in FEAT:
        row[f"sensor_{f}"] = round(sensors.get(f, 0), 4)
    for c in COMPONENTS:
        row[f"fp_{c}"]     = round(get_fp(r, c) * 100, 2)
        row[f"hs_{c}"]     = round(get_hs(r, c) * 100, 2)
        row[f"rl_{c}"]     = get_rl(r, c)
        row[f"pred_{c}"]   = get_pred(r, c)
        row[f"status_{c}"] = get_status(r, c)
    row["overall_health"] = round(r["overall_health_score"] * 100, 2)
    row["recommendation"] = r["recommendation"]
    all_rows.append(row)

# 1. Template bawaan
TEMPLATES = {
    "Normal (Ok-P50)": {
        "rpm": 1500.0, "motor_power": 5874.4252, "noise_db": 51.5149,
        "outlet_pressure_bar": 4.018, "air_flow": 600.1478, "outlet_temp": 119.3749,
        "wpump_outlet_press": 2.8, "water_flow": 55.8112, "gaccz": 3.5898, "haccz": 3.3219,
    },
    "Maintenance Rendah (Noisy-P25)": {
        "rpm": 989.25, "motor_power": 3567.173, "noise_db": 51.4391,
        "outlet_pressure_bar": 2.4125, "air_flow": 594.9976, "outlet_temp": 100.1398,
        "wpump_outlet_press": 2.3454, "water_flow": 58.1976, "gaccz": 2.5713, "haccz": 2.7966,
    },
    "Maintenance Sedang (Noisy-P50)": {
        "rpm": 1499.5, "motor_power": 6138.0733, "noise_db": 58.2337,
        "outlet_pressure_bar": 4.096, "air_flow": 902.107, "outlet_temp": 112.353,
        "wpump_outlet_press": 2.634, "water_flow": 58.4681, "gaccz": 3.6527, "haccz": 3.3943,
    },
    "Maintenance Tinggi (Noisy-P75)": {
        "rpm": 2010.25, "motor_power": 9405.3371, "noise_db": 66.6373,
        "outlet_pressure_bar": 5.7722, "air_flow": 1215.3731, "outlet_temp": 125.5074,
        "wpump_outlet_press": 3.0428, "water_flow": 58.6898, "gaccz": 5.3003, "haccz": 4.1704,
    },
    "Maintenance Kritis (Noisy-P90)": {
        "rpm": 2497.1, "motor_power": 14309.501, "noise_db": 70.3081,
        "outlet_pressure_bar": 6.8842, "air_flow": 1505.4202, "outlet_temp": 137.3772,
        "wpump_outlet_press": 3.3735, "water_flow": 58.8931, "gaccz": 6.7893, "haccz": 4.9894,
    },
}

for name, sensors in TEMPLATES.items():
    all_rows.append({"skenario": name, "sumber": "template", **{}})
    all_rows.pop()  # sementara
    add_row(name, sensors)
    all_rows[-1]["sumber"] = "Template"

# 2. Semua baris dari dataset (sample 200 baris)
print("  Baca dataset...")
sample_df = df_raw.sample(min(200, len(df_raw)), random_state=42)
for i, (_, row) in enumerate(sample_df.iterrows()):
    sensors = {f: float(row[f]) for f in FEAT if f in row and pd.notna(row[f])}
    if len(sensors) < len(FEAT):
        continue
    # label asli dari dataset
    labels = {c: str(row[c]) if c in row else "?" for c in COMPONENTS}
    label_str = " | ".join(f"{COMP_NAMES[c]}:{labels[c]}" for c in COMPONENTS)
    add_row(f"Dataset #{i+1}", sensors)
    all_rows[-1]["sumber"] = "Dataset"
    all_rows[-1]["label_asli"] = label_str

# 3. Random combinations
print("  Random sampling...")
for i in range(300):
    sensors = {}
    for f in FEAT:
        lo = sensor_stats[f]["min"]
        hi = sensor_stats[f]["max"]
        sensors[f] = lo + random.random() * (hi - lo)
    add_row(f"Random #{i+1}", sensors)
    all_rows[-1]["sumber"] = "Random"

# 4. Sweep sensor tunggal (tiap sensor 10 step)
print("  Sweep sensor...")
baseline = {f: sensor_stats[f]["p50"] for f in FEAT}
for sweep_f in FEAT:
    lo = sensor_stats[sweep_f]["min"]
    hi = sensor_stats[sweep_f]["max"]
    for step in range(11):
        val = lo + (hi - lo) * step / 10
        sensors = {**baseline, sweep_f: val}
        add_row(f"Sweep {sweep_f} step{step}", sensors)
        all_rows[-1]["sumber"] = f"Sweep:{sweep_f}"

print(f"  Total baris: {len(all_rows)}")

# ── Buat DataFrame ─────────────────────────────────────────────────────────────
df = pd.DataFrame(all_rows)

# Kolom urutan yang diinginkan
sensor_cols  = [f"sensor_{f}" for f in FEAT]
comp_groups  = []
for c in COMPONENTS:
    comp_groups += [f"pred_{c}", f"status_{c}", f"fp_{c}", f"hs_{c}", f"rl_{c}"]

ordered_cols = (
    ["skenario", "sumber"]
    + sensor_cols
    + comp_groups
    + ["overall_health", "recommendation"]
)
if "label_asli" in df.columns:
    ordered_cols.insert(2, "label_asli")

df = df.reindex(columns=[c for c in ordered_cols if c in df.columns]).fillna("")

# ── Header label lebih mudah dibaca ───────────────────────────────────────────
SENSOR_NICE = {
    "sensor_rpm": "RPM",
    "sensor_motor_power": "Motor Power (W)",
    "sensor_noise_db": "Noise (dB)",
    "sensor_outlet_pressure_bar": "Pressure (bar)",
    "sensor_air_flow": "Air Flow (m3/h)",
    "sensor_outlet_temp": "Outlet Temp (C)",
    "sensor_wpump_outlet_press": "WPump Press (bar)",
    "sensor_water_flow": "Water Flow (L/min)",
    "sensor_gaccz": "G-Acc Z (g)",
    "sensor_haccz": "H-Acc Z (g)",
}

COMP_NICE = {}
for c in COMPONENTS:
    n = COMP_NAMES[c]
    COMP_NICE[f"pred_{c}"]   = f"{n} - Prediksi"
    COMP_NICE[f"status_{c}"] = f"{n} - Status"
    COMP_NICE[f"fp_{c}"]     = f"{n} - Fault Prob (%)"
    COMP_NICE[f"hs_{c}"]     = f"{n} - Health Score (%)"
    COMP_NICE[f"rl_{c}"]     = f"{n} - Risk Level"

HEADER_MAP = {
    "skenario": "Skenario",
    "sumber": "Sumber",
    "label_asli": "Label Asli Dataset",
    "overall_health": "Overall Health Score (%)",
    "recommendation": "Rekomendasi",
    **SENSOR_NICE,
    **COMP_NICE,
}

display_headers = [HEADER_MAP.get(c, c) for c in df.columns]

# ── Buat Excel ─────────────────────────────────────────────────────────────────
print("Membuat Excel...")

wb = Workbook()
ws = wb.active
ws.title = "Hasil Prediksi"

# ── Warna ──────────────────────────────────────────────────────────────────────
C_HEADER_SENSOR  = "1F4E79"  # biru tua - sensor
C_HEADER_BRG     = "7030A0"  # ungu - bearing
C_HEADER_WPU     = "00548C"  # biru - wpump
C_HEADER_RAD     = "375623"  # hijau tua - radiator
C_HEADER_EXV     = "833C00"  # cokelat - exvalve
C_HEADER_META    = "404040"  # abu - metadata
C_HEADER_OVERALL = "B22222"  # merah - overall
C_WHITE          = "FFFFFF"
C_LIGHT_BRG      = "E2CFEF"
C_LIGHT_WPU      = "CFE2FF"
C_LIGHT_RAD      = "D6E4CE"
C_LIGHT_EXV      = "FAE0C8"
C_LIGHT_OVERALL  = "FFD7D7"
C_SENSOR_BG      = "DEEAF1"
C_META_BG        = "F2F2F2"

COMP_COLORS = {
    "bearings":  (C_HEADER_BRG, C_LIGHT_BRG),
    "wpump":     (C_HEADER_WPU, C_LIGHT_WPU),
    "radiator":  (C_HEADER_RAD, C_LIGHT_RAD),
    "exvalve":   (C_HEADER_EXV, C_LIGHT_EXV),
}

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def header_font(color="FFFFFF", bold=True, size=10):
    return Font(bold=bold, color=color, size=size)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def vcenter():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── Row 1: Group header ────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 22
ws.row_dimensions[2].height = 36

group_spans = []  # (start_col, end_col, label, color)

col_idx = {}
for i, c in enumerate(df.columns, 1):
    col_idx[c] = i

# Meta
meta_cols = [c for c in df.columns if c in ("skenario", "sumber", "label_asli")]
if meta_cols:
    group_spans.append((col_idx[meta_cols[0]], col_idx[meta_cols[-1]], "Informasi", C_HEADER_META))

# Sensor
s_cols = [c for c in df.columns if c.startswith("sensor_")]
if s_cols:
    group_spans.append((col_idx[s_cols[0]], col_idx[s_cols[-1]], "Nilai Sensor", C_HEADER_SENSOR))

# Per komponen
for comp in COMPONENTS:
    c_cols = [c for c in df.columns if c.endswith(f"_{comp}")]
    if c_cols:
        hc, _ = COMP_COLORS[comp]
        group_spans.append((col_idx[c_cols[0]], col_idx[c_cols[-1]], COMP_NAMES[comp], hc))

# Overall
ov_cols = [c for c in df.columns if c in ("overall_health", "recommendation")]
if ov_cols:
    group_spans.append((col_idx[ov_cols[0]], col_idx[ov_cols[-1]], "Hasil Keseluruhan", C_HEADER_OVERALL))

for start, end, label, color in group_spans:
    cell = ws.cell(row=1, column=start, value=label)
    cell.fill = fill(color)
    cell.font = header_font()
    cell.alignment = center()
    cell.border = thin_border()
    if end > start:
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)

# ── Row 2: Column headers ──────────────────────────────────────────────────────
for i, (col, header) in enumerate(zip(df.columns, display_headers), 1):
    cell = ws.cell(row=2, column=i, value=header)
    cell.border = thin_border()
    cell.alignment = center()

    # Warna header sesuai group
    if col in ("skenario", "sumber", "label_asli"):
        cell.fill = fill(C_META_BG)
        cell.font = header_font(color="000000", size=9)
    elif col.startswith("sensor_"):
        cell.fill = fill(C_SENSOR_BG)
        cell.font = header_font(color=C_HEADER_SENSOR, size=9)
    elif col in ("overall_health", "recommendation"):
        cell.fill = fill(C_LIGHT_OVERALL)
        cell.font = header_font(color=C_HEADER_OVERALL, size=9)
    else:
        for comp in COMPONENTS:
            if col.endswith(f"_{comp}"):
                hc, lc = COMP_COLORS[comp]
                cell.fill = fill(lc)
                cell.font = header_font(color=hc, size=9)
                break

# ── Data rows ─────────────────────────────────────────────────────────────────
def fp_fill_color(fp_val):
    """Gradient: hijau (0%) → kuning (50%) → merah (100%)"""
    if fp_val is None or fp_val == "":
        return None
    try:
        v = float(fp_val) / 100
    except:
        return None
    if v < 0.1:
        return "C6EFCE"   # hijau muda
    elif v < 0.3:
        return "92D050"   # hijau
    elif v < 0.5:
        return "FFEB9C"   # kuning
    elif v < 0.8:
        return "FFC7CE"   # merah muda
    else:
        return "FF0000"   # merah

def rl_fill_color(rl_val):
    MAP = {
        "very_low": "C6EFCE",
        "low":      "92D050",
        "medium":   "FFEB9C",
        "high":     "FFC7CE",
        "critical": "FF0000",
    }
    return MAP.get(str(rl_val).lower(), None)

def hs_fill_color(hs_val):
    """Gradient terbalik: merah (0%) → hijau (100%)"""
    if hs_val is None or hs_val == "":
        return None
    try:
        v = float(hs_val) / 100
    except:
        return None
    if v >= 0.9:
        return "C6EFCE"
    elif v >= 0.7:
        return "92D050"
    elif v >= 0.5:
        return "FFEB9C"
    elif v >= 0.3:
        return "FFC7CE"
    else:
        return "FF0000"

ROW_OFFSET = 3
for row_i, data_row in enumerate(df.itertuples(index=False), ROW_OFFSET):
    row_data = dict(zip(df.columns, data_row))
    alt = (row_i - ROW_OFFSET) % 2 == 1
    ws.row_dimensions[row_i].height = 15

    for col_i, col_name in enumerate(df.columns, 1):
        val = row_data.get(col_name, "")
        cell = ws.cell(row=row_i, column=col_i, value=val if val != "" else None)
        cell.border = thin_border()
        cell.alignment = vcenter()
        cell.font = Font(size=9)

        # Alignment angka ke kanan
        if isinstance(val, (int, float)):
            cell.alignment = Alignment(horizontal="right", vertical="center")

        # Warna background berdasarkan tipe kolom
        if col_name.startswith("fp_"):
            c = fp_fill_color(val)
            if c:
                cell.fill = fill(c)
                if float(val or 0) >= 80:
                    cell.font = Font(bold=True, color="FFFFFF", size=9)
        elif col_name.startswith("hs_"):
            c = hs_fill_color(val)
            if c:
                cell.fill = fill(c)
        elif col_name.startswith("rl_"):
            c = rl_fill_color(val)
            if c:
                cell.fill = fill(c)
        elif col_name.startswith("pred_") or col_name.startswith("status_"):
            if str(val).lower() in ("noisy", "dirty", "fault detected"):
                cell.fill = fill("FFC7CE")
                cell.font = Font(bold=True, color="9C0006", size=9)
            elif str(val).lower() in ("ok", "clean", "normal"):
                cell.fill = fill("C6EFCE")
                cell.font = Font(color="276221", size=9)
        elif col_name == "overall_health":
            c = hs_fill_color(val)
            if c:
                cell.fill = fill(c)
                cell.font = Font(bold=True, size=9)
        elif col_name.startswith("sensor_"):
            if alt:
                cell.fill = fill("EBF3FB")
        elif col_name in ("skenario", "sumber", "label_asli"):
            if alt:
                cell.fill = fill("F5F5F5")

# ── Lebar kolom ───────────────────────────────────────────────────────────────
col_widths = {
    "skenario": 28, "sumber": 10, "label_asli": 38,
    "overall_health": 14, "recommendation": 55,
}
default_sensor = 15
default_comp   = 12

for i, col_name in enumerate(df.columns, 1):
    ltr = get_column_letter(i)
    if col_name in col_widths:
        ws.column_dimensions[ltr].width = col_widths[col_name]
    elif col_name.startswith("sensor_"):
        ws.column_dimensions[ltr].width = default_sensor
    elif col_name in ("rl_bearings", "rl_wpump", "rl_radiator", "rl_exvalve"):
        ws.column_dimensions[ltr].width = 13
    elif col_name.startswith(("pred_", "status_")):
        ws.column_dimensions[ltr].width = 14
    else:
        ws.column_dimensions[ltr].width = default_comp

# ── Freeze header ──────────────────────────────────────────────────────────────
ws.freeze_panes = "C3"

# ── Auto-filter ───────────────────────────────────────────────────────────────
ws.auto_filter.ref = f"A2:{get_column_letter(len(df.columns))}2"

# ── Sheet 2: Ringkasan statistik per sumber ───────────────────────────────────
ws2 = wb.create_sheet("Ringkasan")
ws2.sheet_view.showGridLines = True

summary_headers = (
    ["Sumber", "Jumlah Baris"]
    + [f"{COMP_NAMES[c]} - Avg FP (%)" for c in COMPONENTS]
    + [f"{COMP_NAMES[c]} - % Baris Fault (fp≥50%)" for c in COMPONENTS]
    + ["Avg Overall Health (%)"]
)
for i, h in enumerate(summary_headers, 1):
    cell = ws2.cell(row=1, column=i, value=h)
    cell.fill = fill(C_HEADER_META)
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.alignment = center()
    cell.border = thin_border()
    ws2.column_dimensions[get_column_letter(i)].width = max(22, len(h) * 0.9)

for row_i, (src, grp) in enumerate(df.groupby("sumber"), 2):
    summary = [src, len(grp)]
    for c in COMPONENTS:
        fc = f"fp_{c}"
        avg_fp = grp[fc].mean() if fc in grp else 0
        summary.append(round(float(avg_fp), 2))
    for c in COMPONENTS:
        fc = f"fp_{c}"
        pct_fault = (grp[fc] >= 50).mean() * 100 if fc in grp else 0
        summary.append(round(float(pct_fault), 1))
    oh = grp["overall_health"].mean() if "overall_health" in grp else 0
    summary.append(round(float(oh), 2))

    for i, v in enumerate(summary, 1):
        cell = ws2.cell(row=row_i, column=i, value=v)
        cell.border = thin_border()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(size=10)
        if row_i % 2 == 0:
            cell.fill = fill("F2F2F2")

ws2.freeze_panes = "A2"

# ── Sheet 3: Contoh per level risiko ──────────────────────────────────────────
ws3 = wb.create_sheet("Contoh per Level Risiko")

risk_rows = []
levels = [
    ("Sangat Rendah (fp<10%)",  lambda fp: fp < 10),
    ("Rendah (10-30%)",         lambda fp: 10 <= fp < 30),
    ("Sedang (30-50%)",         lambda fp: 30 <= fp < 50),
    ("Tinggi (50-80%)",         lambda fp: 50 <= fp < 80),
    ("Kritis (fp>=80%)",        lambda fp: fp >= 80),
]

for comp in COMPONENTS:
    fc = f"fp_{comp}"
    for level_label, cond in levels:
        matches = df[df[fc].apply(lambda x: cond(float(x)) if x != "" else False)]
        if len(matches) == 0:
            continue
        sample = matches.head(3)
        for _, row in sample.iterrows():
            entry = {
                "komponen": COMP_NAMES[comp],
                "level": level_label,
                "skenario": row["skenario"],
                "sumber": row["sumber"],
            }
            for f in FEAT:
                entry[f"sensor_{f}"] = row.get(f"sensor_{f}", "")
            for c in COMPONENTS:
                entry[f"fp_{c}"]   = row.get(f"fp_{c}", "")
                entry[f"hs_{c}"]   = row.get(f"hs_{c}", "")
                entry[f"rl_{c}"]   = row.get(f"rl_{c}", "")
                entry[f"pred_{c}"] = row.get(f"pred_{c}", "")
            entry["overall_health"] = row.get("overall_health", "")
            risk_rows.append(entry)

if risk_rows:
    df_risk = pd.DataFrame(risk_rows)
    risk_headers = list(df_risk.columns)
    risk_nice = {
        "komponen": "Komponen", "level": "Level Risiko",
        "skenario": "Skenario", "sumber": "Sumber",
        **SENSOR_NICE,
        **{f"fp_{c}": f"{COMP_NAMES[c]} FP(%)" for c in COMPONENTS},
        **{f"hs_{c}": f"{COMP_NAMES[c]} HS(%)" for c in COMPONENTS},
        **{f"rl_{c}": f"{COMP_NAMES[c]} Risk" for c in COMPONENTS},
        **{f"pred_{c}": f"{COMP_NAMES[c]} Prediksi" for c in COMPONENTS},
        "overall_health": "Overall Health(%)",
    }
    for i, h in enumerate(risk_headers, 1):
        cell = ws3.cell(row=1, column=i, value=risk_nice.get(h, h))
        cell.fill = fill(C_HEADER_META)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.alignment = center()
        cell.border = thin_border()
        ws3.column_dimensions[get_column_letter(i)].width = 14

    ws3.column_dimensions["A"].width = 16
    ws3.column_dimensions["B"].width = 22
    ws3.column_dimensions["C"].width = 28
    ws3.column_dimensions["D"].width = 12

    prev_comp = None
    prev_level = None
    for row_i, (_, row) in enumerate(df_risk.iterrows(), 2):
        comp_val  = row.get("komponen", "")
        level_val = row.get("level", "")
        is_new_section = (comp_val != prev_comp or level_val != prev_level)
        prev_comp = comp_val
        prev_level = level_val

        for col_i, h in enumerate(risk_headers, 1):
            val = row.get(h, "")
            cell = ws3.cell(row=row_i, column=col_i, value=val if val != "" else None)
            cell.border = thin_border()
            cell.font = Font(size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if h.startswith("fp_"):
                c = fp_fill_color(val)
                if c:
                    cell.fill = fill(c)
            elif h.startswith("rl_"):
                c = rl_fill_color(val)
                if c:
                    cell.fill = fill(c)
            elif h in ("komponen", "level") and is_new_section:
                cell.fill = fill("D9E1F2")
                cell.font = Font(bold=True, size=9)

    ws3.freeze_panes = "E2"
    ws3.auto_filter.ref = f"A1:{get_column_letter(len(risk_headers))}1"

# ── Simpan ─────────────────────────────────────────────────────────────────────
OUTPUT = os.path.join(os.path.dirname(__file__), "hasil_prediksi_model.xlsx")
wb.save(OUTPUT)
print(f"\nFile Excel disimpan: {OUTPUT}")
print(f"Sheet 1 'Hasil Prediksi': {len(df)} baris data")
print(f"Sheet 2 'Ringkasan': statistik per sumber data")
print(f"Sheet 3 'Contoh per Level Risiko': sampel untuk setiap level FP per komponen")
